"""Scan-to-database reconciliation (legacy: compare_scan_to_database).

Parse physical-rack scan exports (.csv/.xlsx/.xls), de-duplicate repeated scan
files, and reconcile against the active feed. Matching is by tube_code only: a
scanned tube is correct if its code is in the active feed, otherwise it is
flagged "scanned, not in database". Rack/box/position are recorded but not used
to match.
"""
from __future__ import annotations

import csv
import io
import re
from pathlib import Path

from ..normalize import normalize_position

# Header aliases (compared lowercased, non-alphanumeric stripped).
_ALIASES = {
    "rackid": "rackid",
    "rack": "rackid",
    "tubecode": "tubecode",
    "barcode": "tubecode",
    "code": "tubecode",
    "tube": "tubecode",
    "position": "position",
    "pos": "position",
    "locationrow": "row",
    "row": "row",
    "locationcolumn": "col",
    "column": "col",
    "col": "col",
    "aliquotid": "typed_aliquot_id",
    "aliquot": "typed_aliquot_id",
    "typedaliquotid": "typed_aliquot_id",
    "extraid": "typed_extra_id",
    "typedextraid": "typed_extra_id",
    "trackid": "typed_extra_id",
}

INVALID_CODES = {"", "noread", "notube", "na", "none", "empty", "nan"}
_RACK_RE = re.compile(r"(\d+)\s*box\s*(\d+)", re.IGNORECASE)


def _key(h: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(h).lower())


def _alias(h: str) -> str | None:
    return _ALIASES.get(_key(h))


def _read_table(filename: str, data: bytes) -> list[list]:
    ext = Path(filename).suffix.lower()
    if ext == ".csv":
        text = data.decode("utf-8-sig", errors="replace")
        return [r for r in csv.reader(io.StringIO(text))]
    if ext == ".xls":
        import xlrd

        book = xlrd.open_workbook(file_contents=data)
        sh = book.sheet_by_index(0)
        return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    # .xlsx / .xlsm
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _project_box(rackid: str) -> tuple[str, str]:
    m = _RACK_RE.search(str(rackid or ""))
    if m:
        return f"L{int(m.group(1))}", str(int(m.group(2)))
    return "", ""


def parse_scan(filename: str, data: bytes) -> list[dict]:
    """Return normalized scan records with rack/project/box/position metadata."""
    table = _read_table(filename, data)
    if not table:
        return []
    header = table[0]
    col = {}
    for i, h in enumerate(header):
        a = _alias(h)
        if a and a not in col:
            col[a] = i

    records: list[dict] = []
    for row in table[1:]:
        def get(name):
            i = col.get(name)
            return row[i] if i is not None and i < len(row) else None

        raw_code = get("tubecode")
        code = "" if raw_code is None else str(raw_code).strip()
        rackid = get("rackid")
        project, box = _project_box(rackid)
        if get("position") is not None and str(get("position")).strip():
            position = normalize_position(get("position"))
        else:
            r, c = get("row"), get("col")
            position = normalize_position(f"{str(r or '').strip()}{str(c or '').strip()}")
        records.append(
            {
                "tube_code": code,
                "project": project,
                "box": box,
                "rack_id": "" if rackid is None else str(rackid).strip(),
                "position": position or "",
                "typed_aliquot_id": "" if get("typed_aliquot_id") is None else str(get("typed_aliquot_id")).strip(),
                "typed_extra_id": "" if get("typed_extra_id") is None else str(get("typed_extra_id")).strip(),
                "source": filename,
            }
        )
    return records


def is_valid_code(code: str) -> bool:
    return _key(code) not in INVALID_CODES


def dedup_files(by_file: dict[str, list[dict]]) -> dict:
    """Group files whose readable codes overlap >= 80%; keep the richest per group."""
    files = list(by_file)
    code_sets = {
        f: {r["tube_code"].upper() for r in recs if is_valid_code(r["tube_code"])}
        for f, recs in by_file.items()
    }

    parent = {f: f for f in files}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    for i in range(len(files)):
        for j in range(i + 1, len(files)):
            a, b = files[i], files[j]
            sa, sb = code_sets[a], code_sets[b]
            if sa and sb:
                overlap = len(sa & sb) / min(len(sa), len(sb))
                if overlap >= 0.8:
                    parent[find(a)] = find(b)

    groups: dict[str, list[str]] = {}
    for f in files:
        groups.setdefault(find(f), []).append(f)

    kept, summary = [], []
    for members in groups.values():
        best = max(members, key=lambda f: len(code_sets[f]))
        for f in members:
            keep = f == best
            if keep:
                kept.append(f)
            summary.append(
                {
                    "file": f,
                    "readable": len(code_sets[f]),
                    "status": "kept" if keep else "drop_duplicate",
                }
            )
    return {"kept": kept, "summary": summary}


def reconcile(db_sheet: dict, records: list[dict]) -> dict:
    """Categorize scan records against the database by code and position."""
    cols = db_sheet["columns"]
    idx = {c: i for i, c in enumerate(cols)}
    cb_i = idx.get("cryobank")
    proj_i = idx.get("project")
    box_i = idx.get("box")
    pos_i = idx.get("sample_pos")
    record_i = idx.get("record_id")

    def cell(row: list, i: int | None):
        return row[i] if i is not None and i < len(row) else None

    def text(value) -> str:
        return "" if value is None else str(value).strip()

    def code_text(value) -> str:
        return text(value).upper()

    def db_slot(row: list) -> tuple[str, str, str]:
        return (
            text(cell(row, proj_i)),
            text(cell(row, box_i)),
            normalize_position(cell(row, pos_i)) or "",
        )

    def row_to_dict(row: list) -> dict:
        return {column: cell(row, i) for i, column in enumerate(cols)}

    def slot_key_from_scan(rec: dict) -> tuple[str, str, str]:
        return (
            text(rec.get("project")),
            text(rec.get("box")),
            normalize_position(rec.get("position")) or "",
        )

    def choose_db_row(candidates: list[list], rec: dict) -> list | None:
        if not candidates:
            return None
        project = text(rec.get("project"))
        box = text(rec.get("box"))
        position = normalize_position(rec.get("position")) or ""
        ranked = sorted(
            candidates,
            key=lambda row: (
                0 if text(cell(row, proj_i)) == project else 1,
                0 if text(cell(row, box_i)) == box else 1,
                0 if (normalize_position(cell(row, pos_i)) or "") == position else 1,
                text(cell(row, record_i)),
            ),
        )
        return ranked[0]

    db_by_code: dict[str, list[list]] = {}
    db_by_slot: dict[tuple[str, str, str], list[list]] = {}
    for row in db_sheet["rows"]:
        slot = db_slot(row)
        db_by_slot.setdefault(slot, []).append(row)
        code = code_text(cell(row, cb_i))
        if is_valid_code(code):
            db_by_code.setdefault(code, []).append(row)

    out = {
        "scan_not_in_database": [],
        "wrong_location": [],
        "database_not_in_scan": [],
        "position_conflicts": [],
        "duplicate_scan_tubecodes": [],
        "correct_matches": 0,
    }

    seen_codes: dict[str, int] = {}
    scanned_valid_codes: set[str] = set()
    observed_slots: dict[tuple[str, str, str], list[dict]] = {}
    observed_boxes: set[tuple[str, str]] = set()
    matched_db_record_ids: set[str] = set()

    for rec in records:
        slot = slot_key_from_scan(rec)
        observed_slots.setdefault(slot, []).append(rec)
        observed_boxes.add((text(rec.get("project")), text(rec.get("box"))))

        code = code_text(rec.get("tube_code"))
        if not is_valid_code(code):
            continue

        scanned_valid_codes.add(code)
        seen_codes[code] = seen_codes.get(code, 0) + 1
        slot_rows = db_by_slot.get(slot, [])
        code_rows = db_by_code.get(code, [])
        exact = next((row for row in code_rows if db_slot(row) == slot), None)

        if exact is not None:
            out["correct_matches"] += 1
            matched_db_record_ids.add(text(cell(exact, record_i)))
            continue

        if code_rows:
            chosen = choose_db_row(code_rows, rec)
            if chosen is not None:
                matched_db_record_ids.add(text(cell(chosen, record_i)))
                out["wrong_location"].append(
                    {
                        **rec,
                        "record_id": cell(chosen, record_i),
                        "expected_project": cell(chosen, proj_i),
                        "expected_box": cell(chosen, box_i),
                        "expected_position": cell(chosen, pos_i),
                        "expected_cryobank": cell(chosen, cb_i),
                    }
                )
            if slot_rows:
                for row in slot_rows:
                    row_code = code_text(cell(row, cb_i))
                    if is_valid_code(row_code) and row_code != code:
                        out["position_conflicts"].append(
                            {
                                **rec,
                                "record_id": cell(row, record_i),
                                "expected_project": cell(row, proj_i),
                                "expected_box": cell(row, box_i),
                                "expected_position": cell(row, pos_i),
                                "expected_cryobank": cell(row, cb_i),
                            }
                        )
        elif slot_rows:
            for row in slot_rows:
                row_code = code_text(cell(row, cb_i))
                if is_valid_code(row_code):
                    out["position_conflicts"].append(
                        {
                            **rec,
                            "record_id": cell(row, record_i),
                            "expected_project": cell(row, proj_i),
                            "expected_box": cell(row, box_i),
                            "expected_position": cell(row, pos_i),
                            "expected_cryobank": cell(row, cb_i),
                        }
                    )
        else:
            out["scan_not_in_database"].append(rec)

    for slot, rows_at_slot in db_by_slot.items():
        if (slot[0], slot[1]) not in observed_boxes:
            continue
        for row in rows_at_slot:
            row_code = code_text(cell(row, cb_i))
            if not is_valid_code(row_code):
                continue
            if row_code not in scanned_valid_codes:
                out["database_not_in_scan"].append(row_to_dict(row))

    out["duplicate_scan_tubecodes"] = [
        {"tube_code": c, "count": n} for c, n in seen_codes.items() if n > 1
    ]
    return out


def build_missing_box_review(db_sheet: dict, missing_records: list[dict]) -> list[dict]:
    """Return DB rows at the scanned slot, plus scanned-tube helpers.

    For each scanned tube not found by code, show the database row(s) already at
    that exact project + box + position. If the slot is empty in the database,
    return one blank database row with the scanned helpers appended.
    """
    columns = db_sheet["columns"]
    idx = {c: i for i, c in enumerate(columns)}
    proj_i, box_i, pos_i = idx.get("project"), idx.get("box"), idx.get("sample_pos")

    def db_row_dict(row: list) -> dict:
        return {column: row[i] if i < len(row) else None for i, column in enumerate(columns)}

    out: list[dict] = []
    for rec_i, rec in enumerate(missing_records):
        rec_project = str(rec.get("project", "")).strip()
        rec_box = str(rec.get("box", "")).strip()
        rec_pos = normalize_position(rec.get("position")) or ""
        candidates = [
            row
            for row in db_sheet["rows"]
            if proj_i is not None
            and box_i is not None
            and pos_i is not None
            and str(row[proj_i]).strip() == rec_project
            and str(row[box_i]).strip() == rec_box
            and (normalize_position(row[pos_i]) or "") == rec_pos
        ]
        if not candidates:
            blank = {column: None for column in columns}
            out.append(
                {
                    **blank,
                    "review_id": f"missing-{rec_i}-0",
                    "scanned_tube_code": rec.get("tube_code"),
                    "scanned_project": rec.get("project"),
                    "scanned_box": rec.get("box"),
                    "scanned_position": rec.get("position"),
                    "scanned_source": rec.get("source"),
                }
            )
            continue
        for row_i, row in enumerate(candidates):
            out.append(
                {
                    **db_row_dict(row),
                    "review_id": f"missing-{rec_i}-{row_i}",
                    "scanned_tube_code": rec.get("tube_code"),
                    "scanned_project": rec.get("project"),
                    "scanned_box": rec.get("box"),
                    "scanned_position": rec.get("position"),
                    "scanned_source": rec.get("source"),
                }
            )
    return out


def build_wrong_location_review(db_sheet: dict, wrong_rows: list[dict]) -> list[dict]:
    """Return full-database rows with scanned-location helpers for wrong-location fixes."""
    columns = db_sheet["columns"]
    idx = {c: i for i, c in enumerate(columns)}
    record_i = idx.get("record_id")

    def db_row_dict(row: list) -> dict:
        return {column: row[i] if i < len(row) else None for i, column in enumerate(columns)}

    by_record = {}
    for row in db_sheet["rows"]:
        if record_i is not None and record_i < len(row):
            by_record[str(row[record_i])] = row

    out: list[dict] = []
    for i, item in enumerate(wrong_rows):
        row = by_record.get(str(item.get("record_id", "")))
        if row is None:
            continue
        out.append(
            {
                **db_row_dict(row),
                "review_id": f"wrong-{i}",
                "scanned_box": item.get("box"),
                "scanned_position": item.get("position"),
                "scanned_source": item.get("source"),
            }
        )
    return out
