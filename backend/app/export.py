"""Build styled .xlsx workbooks from rows for download (legacy tools deliver xlsx)."""
from __future__ import annotations

import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_INVALID_SHEET = re.compile(r"[\[\]:*?/\\]")
_HEADER_FILL = PatternFill("solid", fgColor="0E8ED6")  # Kolaboratory sky blue
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _safe_sheet_name(name: str) -> str:
    cleaned = _INVALID_SHEET.sub(" ", name or "").strip()
    return (cleaned or "Export")[:31]


def build_xlsx(columns: list[str], rows: list[list], sheet_name: str = "Export") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(sheet_name)

    ws.append(columns)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    for row in rows:
        ws.append(list(row))

    ws.freeze_panes = "A2"
    if columns:
        ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"
        ws.auto_filter.ref = ref

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def build_multi_xlsx(sheets: dict[str, tuple[list[str], list[list]]]) -> bytes:
    """Build a workbook with one sheet per (name -> (columns, rows))."""
    wb = Workbook()
    wb.remove(wb.active)
    for name, (columns, rows) in sheets.items():
        ws = wb.create_sheet(_safe_sheet_name(name))
        ws.append(columns or ["(empty)"])
        _style_header(ws)
        for row in rows:
            ws.append(list(row))
        ws.freeze_panes = "A2"
    if not wb.sheetnames:
        wb.create_sheet("Empty")
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def dicts_to_table(items: list[dict]) -> tuple[list[str], list[list]]:
    """Flatten a list of dicts to (columns, rows) using the union of keys."""
    columns: list[str] = []
    for it in items:
        for k in it:
            if k not in columns:
                columns.append(k)
    rows = [[it.get(c) for c in columns] for it in items]
    return columns, rows
