"""Parse uploaded spreadsheets into normalized, JSON-serializable sheet data.

Output shape (one dict per sheet)::

    {
        "name": str,
        "columns": list[str],          # header row, normalized
        "rows": list[list],            # data rows (capped to MAX_ROWS_PER_SHEET)
        "rowCount": int,               # TRUE total of non-empty data rows
        "truncated": bool,             # True when rowCount > len(rows)
        "schemaValid": bool,           # matches a registered schema exactly
        "issues": list[dict],          # schema validation issues
    }

Normalization rules (spec §4.4): trim string whitespace, blank cells -> None,
datetime -> ISO string, drop fully-empty rows. Rows are capped for display while
the true total is always reported, so nothing is silently dropped.
"""
from __future__ import annotations

import csv
import datetime as _dt
from pathlib import Path
from typing import Any, Iterable, Iterator

from openpyxl import load_workbook

from .schemas import registry


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped != "" else None
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.isoformat()
    return value


def _is_empty_row(row: list[Any]) -> bool:
    return all(cell is None for cell in row)


def _clean_header(header: Iterable[Any]) -> list[str]:
    columns = [str(h).strip() if h is not None else "" for h in header]
    while columns and columns[-1] == "":
        columns.pop()
    return columns


def _build_sheet(
    name: str,
    header: Iterable[Any],
    row_iter: Iterator[Iterable[Any]],
    max_rows: int | None,
) -> dict:
    """Build a normalized sheet. ``max_rows=None`` keeps every row (used for
    storage); a number caps the kept rows (used by tests / previews)."""
    columns = _clean_header(header)
    width = len(columns)

    rows: list[list[Any]] = []
    row_count = 0
    for raw in row_iter:
        raw_list = list(raw)
        norm = [_normalize_cell(c) for c in raw_list[:width]]
        if len(norm) < width:
            norm += [None] * (width - len(norm))
        if _is_empty_row(norm):
            continue
        row_count += 1
        if max_rows is None or len(rows) < max_rows:
            rows.append(norm)

    match_status, schema_name, issues = registry.classify(columns)

    return {
        "name": name,
        "columns": columns,
        "rows": rows,
        "rowCount": row_count,
        "truncated": row_count > len(rows),
        "match": match_status,  # 'matched' | 'partial' | 'other'
        "schema": schema_name,
        "schemaValid": match_status == "matched",
        "issues": issues,
    }


def parse_xlsx(path: str | Path, max_rows: int | None = None) -> list[dict]:
    cap = max_rows  # None = keep all rows
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    sheets: list[dict] = []
    try:
        for ws in wb.worksheets:
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = list(next(rows_iter))
            except StopIteration:
                sheets.append(_build_sheet(ws.title, [], iter(()), cap))
                continue
            sheets.append(_build_sheet(ws.title, header, rows_iter, cap))
    finally:
        wb.close()
    return sheets


def parse_csv(path: str | Path, max_rows: int | None = None) -> list[dict]:
    cap = max_rows  # None = keep all rows
    with open(path, "r", newline="", encoding="utf-8-sig") as f:
        reader = list(csv.reader(f))
    if not reader:
        return [_build_sheet("CSV", [], iter(()), cap)]
    header, data_rows = reader[0], reader[1:]
    return [_build_sheet("CSV", header, iter(data_rows), cap)]


def parse_file(path: str | Path, ext: str, max_rows: int | None = None) -> list[dict]:
    ext = ext.lower()
    if ext == ".csv":
        return parse_csv(path, max_rows)
    return parse_xlsx(path, max_rows)
